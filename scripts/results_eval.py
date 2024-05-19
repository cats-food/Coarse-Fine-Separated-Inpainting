import numpy as np
import matplotlib.pyplot as plt
from glob import glob
from ntpath import basename
from scipy.misc import imread
from skimage.measure import compare_ssim, compare_psnr
import os
from openpyxl import Workbook, load_workbook

'''
该脚本对测试样本进行质量评估，并将评估结果保存到excel表格中。
表格第1列：输出图片名
表格第2列：相对MAE。计算方法是 ||pred-gt|| / ||gt||， 其中||x||是1范数
表格第3列：PSNR
表格第4列：SSIM （窗口大小11）

'''

path_gt = '/home/ysy/batch_test/psv/__GT__'  # 真值图片路径（必须是图片，不得是flist）
path_pred = '/home/ysy/batch_test/psv/FCA5.5(noAtt)_630k_5k'  # 预测图片路径（必须是图片，不得是flist）
xlsx = '/home/ysy/batch_test/psv/psv.xlsx'  # 存放评估数值的excel表格
workseetName = 'noAtt_630k_5k'  # CA  GL  EC  PIC  RFR
debug = 0

assert os.path.exists(xlsx), 'ysy: invalid xlsx path!'
workbook = Workbook()  # 实例化一个工作簿对象
workbook = load_workbook(filename=xlsx)  # workbook加载对应excel文件
#worksheet = workbook[workbook.get_sheet_names()[0]]  # 采用键值索引方式加载excel里面指定的sheet（默认是Sheet1）
worksheet = workbook[workseetName]  # 采用键值索引方式加载excel里面指定的sheet（默认是Sheet1）
worksheet.cell(1, 1).value ='Name'
worksheet.cell(1, 2).value ='MAE'
worksheet.cell(1, 3).value ='PSNR'
worksheet.cell(1, 4).value ='SSIM'


def compare_relative_mae(img_true, img_test):
    img_true = img_true.astype(np.float32)
    img_test = img_test.astype(np.float32)
    return 100*np.sum(np.abs(img_true - img_test)) / np.sum(img_true)  # 乘以100是表示成百分数的形式


psnr_acc = []
ssim_acc = []
mae_acc = []


files_gt = sorted(list(glob(path_gt + '/*.png')))  # +list(glob(path_gt + '/*.jpg')) +
files_pred = sorted(list(glob(path_pred + '/*.png')))  # +list(glob(path_pred + '/*.jpg')) +

if len(files_gt) != len(files_pred):
    raise Exception('ysy: numbers should be the same!')

for i in range(0, len(files_gt)):
    names_gt=basename(str(files_gt[i]))
    names_pred=basename(str(files_pred[i]))

    # if names_gt[:6] != names_pred[:6]:
    #     raise Exception('ysy: mismatched gt-pred pair!')

    img_gt = imread(files_gt[i]).astype(np.float32)
    img_pred = imread(files_pred[i]).astype(np.float32)

    print(f"test pair: gt: {files_gt[i]}         img: {files_pred[i]}")

    if debug != 0:
        plt.subplot('121')
        plt.imshow(img_gt)
        plt.title('gt')
        plt.subplot('122')
        plt.imshow(img_pred)
        plt.title('pred')
        plt.show()

    mae = compare_relative_mae(img_gt, img_pred)
    psnr = compare_psnr(img_gt, img_pred, data_range=255)
    ssim = compare_ssim(img_gt, img_pred, multichannel=True, data_range=255, win_size=11)

    worksheet.cell(i+2, 1).value = str(names_pred)
    worksheet.cell(i+2, 2).value = mae
    worksheet.cell(i+2, 3).value = psnr
    worksheet.cell(i+2, 4).value = ssim

    psnr_acc.append(psnr)
    ssim_acc.append(ssim)
    mae_acc.append(mae)

    if np.mod(i + 1, 10) == 0:
        print(str(i + 1) + ' images processed')

# np.savez('/metrics.npz', psnr=psnr, ssim=ssim, mae=mae, names_gt=names_gt)
# print(
#     "Mean PSNR: %.4f" % round(np.mean(psnr), 4),
#     "PSNR Variance: %.4f" % round(np.var(psnr), 4),
#     "Mean SSIM: %.4f" % round(np.mean(ssim), 4),
#     "SSIM Variance: %.4f" % round(np.var(ssim), 4),
#     "Mean MAE: %.4f" % round(np.mean(mae), 4),
#     "MAE Variance: %.4f" % round(np.var(mae), 4)
# )
workbook.save(xlsx)
step=100  # 每个mask ratio区间有多少个样本。比如step=100代表： 0-0.1， 0.1-0.2， ... 0.4-0.5每个区间都有100个样本。
print("Mean MAE in 5 mask ratio intervals: %.4f %.4f %.4f %.4f %.4f" % (round(np.mean(mae_acc[0:step-1]), 4), round(np.mean(mae_acc[step:2*step-1]), 4), round(np.mean(mae_acc[2*step:3*step-1]), 4),round(np.mean(mae_acc[3*step:4*step-1]), 4), round(np.mean(mae_acc[4*step:5*step-1]), 4)))
print("Mean PSNR in 5 mask ratio intervals: %.4f %.4f %.4f %.4f %.4f" % (round(np.mean(psnr_acc[0:step-1]), 4),round(np.mean(psnr_acc[step:2*step-1]), 4),round(np.mean(psnr_acc[2*step:3*step-1]), 4),round(np.mean(psnr_acc[3*step:4*step-1]), 4),round(np.mean(psnr_acc[4*step:5*step-1]), 4)))
print("Mean SSIM in 5 mask ratio intervals: %.4f %.4f %.4f %.4f %.4f" % (round(np.mean(ssim_acc[0:step-1]), 4),round(np.mean(ssim_acc[step:2*step-1]), 4),round(np.mean(ssim_acc[2*step:3*step-1]), 4),round(np.mean(ssim_acc[3*step:4*step-1]), 4),round(np.mean(ssim_acc[4*step:5*step-1]), 4)))
print('done!')