from openpyxl import Workbook
from openpyxl import load_workbook
import os
import torch
import cv2
import random
import numpy as np
# from main import load_config
from src.engine import Engine
from src.config import Config
##########################################################################
## input arguments  注意：该脚本与config.yml文件完全脱钩，所有参数均在下面设置！
########################################################################
mask_in_path = '../examples/celeba/masks'  # 输入mask路径
img_in_path = '../examples/celeba/images'  # 输入图片路径
img_out_path = '../checkpoints/celeba/tmp'  #700kG1_0-300kG2'  # 输出图片路径
xlsx = '../checkpoints/celeba/x.xlsx'  # 记录结果的excel表格
pth_eval_path = '../checkpoints/celeba/'  # 要验证的pth文件路径

model = 4        # 1(G1 only) | 2(G2 only, less used) | 3(single G1, multi G2) |  4 (multi G1, multi G2)
if model == 1:
    first_pth1_to_eval = 'CoarseModel_G_00180000.pth'  # 从哪个pth开始验证（含）,do not attach the prefix. Set '0' to eval start over.
elif model == 2:
    first_pth2_to_eval = 'RefineModel_G_00350000.pth'  # 从哪个pth开始验证（含）,do not attach the prefix. Set '0' to eval start over.
elif model == 3:
    pth1_path = '../checkpoints/celeba/CoarseModel_G_00340000.pth'  # model不等的时候（即固定inp1批量验证inp2时）填，model=1时该项无�?
    first_pth2_to_eval = 'RefineModel_G_00180000.pth'  # 从哪个pth开始验证（含）,do not attach the prefix. Set '0' to eval start over.
elif model == 4:
    first_pth1_to_eval = 'CoarseModel_G_00340000.pth'  # 从哪个pth开始验证（含）,do not attach the prefix. Set '0' to eval start over.
    first_pth2_to_eval = 'RefineModel_G_00240000.pth'  # 从哪个pth开始验证（含）,do not attach the prefix. Set '0' to eval start over.

############################################################################
# 注意：该脚本与config.yml文件完全脱钩，所有参数均在上面设置！
#############################################################################
assert os.path.exists(mask_in_path), 'ysy: invalid mask_in path!'
assert os.path.exists(img_in_path), 'ysy: invalid img_in path!'
assert os.path.exists(xlsx), 'ysy: invalid xlsx path!'
assert os.path.exists(img_out_path), 'ysy: invalid img_out_path path!'
assert os.path.exists(pth_eval_path), 'ysy: invalid pth_eval_path!'

# if model!=1:
#     assert os.path.exists(pth1_path), 'ysy: you need a specific pth1 !'

workbook = Workbook()  # 实例化一个工作簿对象
workbook = load_workbook(filename=xlsx)  # workbook加载对应excel文件
#worksheet = workbook[workbook.get_sheet_names()[0]]  # 采用键值索引方式加载excel里面指定的sheet（默认是Sheet1�
worksheet = workbook['tmp']  # 采用键值索引方式加载excel里面指定的sheet（默认是Sheet1�


def check_match_of_pth1_and_pth2(pth1_files,pth2_files):
    '''
    :param pth1_files: list of pth1_files
    :param pth2_files: list of pth2_files
    :return: True or False. if two lists are matched, i.e., they have the same length and the same iter lag, then
    returns True. Otherwise returns False
    '''
    if len(pth1_files) != len(pth2_files):
        print('ysy: pth1_files and pth2_files should have the same length!')
        return False
    s = pth1_files[0].find('0')
    lag = int(pth1_files[0][s:s+8]) - int(pth2_files[0][s:s+8])
    for idx in range(0, len(pth1_files)):
        if (int(pth1_files[idx][s:s+8]) - int(pth2_files[idx][s:s+8])) != lag:
            return False
    return True


#config = load_config(2)
config = Config('../config.yml')
os.environ['CUDA_VISIBLE_DEVICES'] = ','.join(str(e) for e in config.GPU)
if torch.cuda.is_available():
    config.DEVICE = torch.device("cuda")
    torch.backends.cudnn.benchmark = True  # cudnn auto-tuner
else:
    config.DEVICE = torch.device("cpu")
cv2.setNumThreads(0)
torch.manual_seed(config.SEED)
torch.cuda.manual_seed_all(config.SEED)
np.random.seed(config.SEED)
random.seed(config.SEED)

config.RESULTS = img_out_path
config.TEST_MASK_FLIST = mask_in_path
config.TEST_FLIST = img_in_path
config.MODE = 2
config.MODEL = model

pth1_files=[]
pth2_files=[]

if config.MODEL == 1:   # multi pth1 need to be put in a list
    for root, dir, files in os.walk(pth_eval_path):
        for file in files:
            if '1_G_' in file and (file >= first_pth1_to_eval) and 'pth' in file:
                pth1_files.append(os.path.join(root,file))
    pth1_files.sort()
elif config.MODEL == 2 or config.MODEL == 3:       # multi pth2 need to be put in a list
    for root, dir, files in os.walk(pth_eval_path):
        for file in files:
            if '2_G_' in file and (file >= first_pth2_to_eval) and 'pth' in file:
                pth2_files.append(os.path.join(root,file))
    pth2_files.sort()
elif config.MODEL == 4:  # multi pth1 and pth2 need to be put in two lists
    for root, dir, files in os.walk(pth_eval_path):
        for file in files:
            if '1_G_' in file and (file >= first_pth1_to_eval) and 'pth' in file:
                pth1_files.append(os.path.join(root,file))
            if '2_G_' in file and (file >= first_pth2_to_eval) and 'pth' in file:
                pth2_files.append(os.path.join(root,file))
    pth1_files.sort()
    pth2_files.sort()

    assert check_match_of_pth1_and_pth2(pth1_files,pth2_files), 'ysy: mismatched pth1 - pth2 files!'  # check match or not

num_pth_files = max(len(pth1_files), len(pth2_files))

for pth_idx in range(0, num_pth_files):
    #config.EdgeModel_G_LOAD_PATH = '.\checkpoints\celeba\EdgeModel_gen.pth'
    #config.RefineModel_G_LOAD_PATH = os.path.join(root, pt_files[pth_idx])
    if config.MODEL == 1: # 验证批量inp1
        config.CoarseModel_G_LOAD_PATH = pth1_files[pth_idx]
        config.RefineModel_G_LOAD_PATH = '--'
    elif config.MODEL == 2:
        config.CoarseModel_G_LOAD_PATH = '--'
        config.RefineModel_G_LOAD_PATH = pth2_files[pth_idx]
    elif config.MODEL == 3:    # 固定一个inp1，验证批量inp2
        config.CoarseModel_G_LOAD_PATH = pth1_path
        config.RefineModel_G_LOAD_PATH = pth2_files[pth_idx]
    elif config.MODEL == 4:
        config.CoarseModel_G_LOAD_PATH = pth1_files[pth_idx]
        config.RefineModel_G_LOAD_PATH = pth2_files[pth_idx]
    # try:
    model = Engine(config)
    model.load()
    test_info = model.test(multi_pth_eval=True)  # test_info[0]: iter1   test_info[1]: iter2    test_info[2]: img idx    test_info[3]:img name    test_info[4]: psnr    test_info[5]: mae
    # except:
    #     test_info = [[config.RefineModel_G_LOAD_PATH[-11:-4], '#nan#','#nan#','#nan#','#nan#']]
    #     print('ysy: error loading: '+config.RefineModel_G_LOAD_PATH)

    worksheet.cell(pth_idx + 2, 1).value = str(test_info[0][0])  # 表格每一行的第一列是pth1的iter?
    worksheet.cell(pth_idx + 2, 2).value = str(test_info[0][1])  # 表格每一行的第一列是pth2的iter?

    for imgIdx in range(0, len(test_info)):
        worksheet.cell(pth_idx + 2, imgIdx + 3).value = test_info[imgIdx][4]  # 把该pth测试的所有样本的psnr写入到excel表格里的第pth_idx+2行中，从第3列开�?


for imgIdx in range(0, len(test_info)):
    worksheet.cell(1, imgIdx + 3).value = test_info[imgIdx][3]  # excel第一行是图片name
worksheet.cell(1, 1).value = 'G1 iter'
worksheet.cell(1, 2).value = 'G2 iter'


workbook.save(xlsx)
print('done')




