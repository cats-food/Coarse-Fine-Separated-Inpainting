# 需要 conda install openpyxl
# 用法，先把mode设置为'export_to_xlsx'，运行后，src pth的字典参数会保存到xlsx的第一列中，
# 然后打开xlsx，复制第一列到第二列，然后修改第二列中的想要重命名的键，完成后保存关闭xlsx。
# 再回到该脚本把mode设置为'update_from_xlsx'即可把更新后的pth保存到dst pth
from openpyxl import Workbook
from openpyxl import load_workbook
import torch

mode = 'export_to_xlsx'
mode = 'update_from_xlsx'

xlsx = 'E:/Desktop/pth_dict.xlsx'
src_pth = "E:\Projects_py_local\Edge-guided Attentive Model v.4\checkpoints\psv\CoarseModel_G_00000000.pth"
dst_pth = 'E:\Projects_py_local\Edge-guided Attentive Model v.4\checkpoints\psv\CoarseModel_G_00000000.pth'

workbook = Workbook()  # 实例化一个工作簿对象
workbook = load_workbook(filename=xlsx)  # workbook加载对应excel文件
worksheet = workbook[workbook.get_sheet_names()[0]]  # 采用键值索引方式加载excel里面指定的sheet（默认是Sheet1）


pt = torch.load(src_pth)

# extract info from pth
if mode == 'export_to_xlsx':
    #workbook.remove_sheet(workbook.get_sheet_names()[0])
    #workbook.create_sheet(workbook.get_sheet_names()[0])
    i = 1
    for k in list(pt['generator'].keys()):
        worksheet.cell(i, 1).value = k
        i += 1
    workbook.save(xlsx)
    print('extract successfully!')

# update info for pth
elif mode == 'update_from_xlsx':
    i = 1
    # pt['generator']['iteration'] = 0  # 修改pt文件的iter
    for old_k in list(pt['generator'].keys()):
        new_k = worksheet.cell(i, 2).value
        # if pt['generator'][new_k].shape != pt['generator'][old_k].shape: #检查值的尺寸是否匹配
        #     raise Exception('ysy: unmatched tensor size for new-old key pair!: %s -- %s' %(new_k,old_k))
        pt['generator'][new_k] = pt['generator'].pop(old_k)
        i += 1
    torch.save(pt, dst_pth)
    print('update successfully!')

else:
    raise Exception('Invalid mode!')
